import pandas as pd
import openpyxl as xl
from openpyxl import Workbook
import datetime as dt

wb = Workbook()

print(pd.__version__)

ws = wb.active

ws['A1'] = 'Hello, world!'
ws['A2'] = dt.datetime.now()

wb.save('test.xlsx')